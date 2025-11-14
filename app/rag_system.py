#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
System RAG dla dokumentów prawnych w języku polskim
Obsługuje PDF, DOCX, XLSX, obrazy (JPG, PNG, BMP)
Wykorzystuje lokalne modele Ollama: Gemma 3:12B do opisu grafik i rozmów o prawie
"""

import os
import sys
import json
import uuid
import logging
from typing import List, Dict, Any, Tuple
from dataclasses import dataclass, field
from pathlib import Path
import shutil
import time

# WYŁĄCZENIE LOGOWANIA PDFMINER NA SAMYM POCZĄTKU
try:
    # Metoda 1: Wyłączenie logowania pdfminer
    import pdfminer
    import pdfminer.psparser
    import pdfminer.pdfinterp
    
    # Ustawienie poziomu logowania na CRITICAL (minimalne logi)
    logging.getLogger("pdfminer").setLevel(logging.CRITICAL)
    logging.getLogger("pdfminer.psparser").setLevel(logging.CRITICAL)
    logging.getLogger("pdfminer.pdfinterp").setLevel(logging.CRITICAL)
    logging.getLogger("pdfminer.pdfdevice").setLevel(logging.CRITICAL)
    logging.getLogger("pdfminer.pdffont").setLevel(logging.CRITICAL)
    logging.getLogger("pdfminer.pdfpage").setLevel(logging.CRITICAL)
    logging.getLogger("pdfminer.pdfdocument").setLevel(logging.CRITICAL)
    
    # Metoda 2: Przekierowanie stdout stderr pdfminer (dodatkowa ochrona)
    import io
    from contextlib import redirect_stdout, redirect_stderr
    
except ImportError:
    pass

# Ekstrakcja danych
import PyPDF2
import pdfplumber
from docx import Document
import openpyxl
from PIL import Image
import pytesseract

# Przetwarzanie języka naturalnego
from sentence_transformers import SentenceTransformer
import chromadb
from chromadb.config import Settings

# Komunikacja z Ollama
import requests

# Filtrowanie powitań
from greeting_filter import GreetingFilter

# Model providers (OpenAI, Ollama)
from model_provider import ModelFactory, ModelProvider

# Hybrydowe wyszukiwanie
from hybrid_search import HybridSearch

# Audit logging
from audit_logger import get_audit_logger

# Device management (GPU/CPU)
from device_manager import DeviceManager

# Web search (intranet/internet)
from web_search import BingSearchProvider, WebScraper, WebSearchCache

# Konfiguracja logowania - bardziej szczegółowa
logging.basicConfig(
    level=logging.INFO,  # Zmienione z DEBUG na INFO
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('rag_system.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

# Dodatkowe wyłączenie logowania pdfminer (jeszcze bardziej agresywne)
os.environ['PDFMINER_LOG_LEVEL'] = 'CRITICAL'

# Konfiguracja ścieżek
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
VECTOR_DB_DIR = BASE_DIR / "vector_db"
TEMP_DIR = BASE_DIR / "temp"
MODELS_DIR = BASE_DIR / "models"
WHISPER_MODELS_DIR = MODELS_DIR / "whisper"
EMBEDDING_MODELS_DIR = MODELS_DIR / "embeddings"
RERANKER_MODELS_DIR = MODELS_DIR / "reranker"

# Tworzenie katalogów jeśli nie istnieją
DATA_DIR.mkdir(exist_ok=True)
VECTOR_DB_DIR.mkdir(exist_ok=True)
TEMP_DIR.mkdir(exist_ok=True)
for directory in [MODELS_DIR, WHISPER_MODELS_DIR, EMBEDDING_MODELS_DIR, RERANKER_MODELS_DIR]:
    directory.mkdir(parents=True, exist_ok=True)

# Konfiguracja modeli
VISION_MODEL = "gemma3:12b"  # Model multimodalny do opisu grafik
LLM_MODEL = "gemma3:12b"     # Model do generowania odpowiedzi

# Plik z sugerowanymi pytaniami
SUGGESTED_QUESTIONS_FILE = BASE_DIR / "suggested_questions.json"

@dataclass
class DocumentChunk:
    """Reprezentacja fragmentu dokumentu"""
    id: str
    content: str
    source_file: str
    page_number: int
    chunk_type: str  # 'text' lub 'image_description'
    element_id: str = ""  # ID elementu w dokumencie (np. numer sekcji)
    embedding: List[float] = field(default_factory=list)

@dataclass
class SourceReference:
    """Reprezentacja odniesienia do źródła"""
    source_file: str
    page_number: int
    element_id: str
    content: str
    distance: float = 0.0
    
    def __post_init__(self):
        """Walidacja po inicjalizacji - ignoruj nieznane argumenty"""
        pass

class DocumentProcessor:
    """Klasa do przetwarzania różnych formatów dokumentów"""
    
    def __init__(self):
        self.supported_formats = {'.pdf', '.docx', '.xlsx', '.jpg', '.jpeg', '.png', '.bmp'}
        logger.info("Inicjalizacja DocumentProcessor")
    
    def process_directory(self, directory_path: str) -> List[DocumentChunk]:
        """Przetwarza wszystkie obsługiwane pliki w katalogu"""
        logger.info(f"Rozpoczynanie przetwarzania katalogu: {directory_path}")
        chunks = []
        path = Path(directory_path)
        
        if not path.exists():
            logger.error(f"Katalog {directory_path} nie istnieje")
            return chunks
            
        # Policz pliki do przetworzenia
        files_to_process = [f for f in path.rglob('*') if f.is_file() and f.suffix.lower() in self.supported_formats]
        logger.info(f"Znaleziono {len(files_to_process)} plików do przetworzenia")
        
        processed_files = 0
        for file_path in files_to_process:
            processed_files += 1
            logger.info(f"[{processed_files}/{len(files_to_process)}] Rozpoczynanie przetwarzania pliku: {file_path}")
            try:
                start_time = time.time()
                file_chunks = self.process_file(file_path)
                processing_time = time.time() - start_time
                logger.info(f"[{processed_files}/{len(files_to_process)}] Zakończono przetwarzanie {file_path} w {processing_time:.2f} sekund. Znaleziono {len(file_chunks)} fragmentów")
                chunks.extend(file_chunks)
            except Exception as e:
                logger.error(f"[{processed_files}/{len(files_to_process)}] Błąd podczas przetwarzania pliku {file_path}: {e}", exc_info=True)
        
        logger.info(f"Zakończono przetwarzanie katalogu. Łącznie fragmentów: {len(chunks)}")
        return chunks
    
    def process_file(self, file_path: Path) -> List[DocumentChunk]:
        """Przetwarza pojedynczy plik zgodnie z jego rozszerzeniem"""
        suffix = file_path.suffix.lower()
        logger.debug(f"Rozpoczynanie przetwarzania pliku {file_path} (typ: {suffix})")
        
        if suffix == '.pdf':
            return self._process_pdf(file_path)
        elif suffix == '.docx':
            return self._process_docx(file_path)
        elif suffix == '.xlsx':
            return self._process_xlsx(file_path)
        elif suffix in ['.jpg', '.jpeg', '.png', '.bmp']:
            return self._process_image(file_path)
        elif suffix in ['.mp3', '.wav', '.flac', '.ogg', '.m4a']:
            return self._process_audio(file_path)
        elif suffix in ['.mp4', '.avi', '.mov', '.mkv', '.webm']:
            return self._process_video(file_path)
        else:
            logger.warning(f"Nieobsługiwany format pliku: {suffix}")
            return []
    
    def _process_pdf(self, file_path: Path) -> List[DocumentChunk]:
        """Przetwarza plik PDF"""
        logger.info(f"Rozpoczynanie przetwarzania PDF: {file_path}")
        chunks = []
        
        try:
            # Wyłączenie logowania pdfminer w kontekście przetwarzania
            pdfminer_logger = logging.getLogger("pdfminer")
            original_level = pdfminer_logger.level
            pdfminer_logger.setLevel(logging.CRITICAL)
            
            with pdfplumber.open(file_path) as pdf:
                logger.info(f"PDF {file_path} ma {len(pdf.pages)} stron")
                for page_num, page in enumerate(pdf.pages, 1):
                    logger.debug(f"Przetwarzanie strony {page_num}")
                    
                    # Ekstrakcja tekstu
                    text = page.extract_text()
                    if text:
                        text_chunks = self._chunk_text(text)
                        logger.debug(f"Znaleziono tekst na stronie {page_num}, podzielono na {len(text_chunks)} fragmentów")
                        for i, chunk in enumerate(text_chunks):
                            chunks.append(DocumentChunk(
                                id=str(uuid.uuid4()),
                                content=chunk,
                                source_file=file_path.name,
                                page_number=page_num,
                                chunk_type='text',
                                element_id=f"tekst_{page_num}_{i+1}"
                            ))
                    
                    # Ekstrakcja obrazów
                    if page.images:
                        logger.info(f"Znaleziono {len(page.images)} grafik na stronie {page_num}")
                        for img_idx, img_obj in enumerate(page.images):
                            try:
                                logger.debug(f"Przetwarzanie grafiki {img_idx+1}/{len(page.images)} na stronie {page_num}")
                                # Zapis obrazu tymczasowo
                                img_data = img_obj['stream'].get_data()
                                img_path = TEMP_DIR / f"temp_img_{uuid.uuid4()}.png"
                                
                                with open(img_path, 'wb') as f:
                                    f.write(img_data)
                                
                                # Opis obrazu
                                description = self._describe_image(img_path)
                                if description:
                                    logger.debug(f"Wygenerowano opis grafiki, długość: {len(description)} znaków")
                                    chunks.append(DocumentChunk(
                                        id=str(uuid.uuid4()),
                                        content=description,
                                        source_file=file_path.name,
                                        page_number=page_num,
                                        chunk_type='image_description',
                                        element_id=f"grafika_{page_num}_{img_idx+1}"
                                    ))
                                
                                # Usunięcie tymczasowego pliku
                                img_path.unlink()
                            except Exception as e:
                                logger.error(f"Błąd podczas przetwarzania obrazu {img_idx+1} na stronie {page_num}: {e}")
            
            # Przywrócenie oryginalnego poziomu logowania
            pdfminer_logger.setLevel(original_level)
            logger.info(f"Zakończono przetwarzanie PDF {file_path}, znaleziono łącznie {len(chunks)} fragmentów")
            
        except Exception as e:
            logger.error(f"Błąd podczas przetwarzania PDF {file_path}: {e}", exc_info=True)
        
        return chunks
    
    def _process_docx(self, file_path: Path) -> List[DocumentChunk]:
        """Przetwarza plik DOCX (tekst + obrazy)"""
        logger.info(f"Rozpoczynanie przetwarzania DOCX: {file_path}")
        chunks = []
        
        try:
            doc = Document(file_path)
            full_text = []
            section_counter = 0
            
            # Wyciągnij tekst
            for i, paragraph in enumerate(doc.paragraphs):
                if paragraph.text.strip():
                    full_text.append(paragraph.text)
                    section_counter += 1
            
            text = '\n'.join(full_text)
            if text:
                text_chunks = self._chunk_text(text)
                logger.debug(f"Znaleziono tekst w DOCX, podzielono na {len(text_chunks)} fragmentów")
                for i, chunk in enumerate(text_chunks):
                    chunks.append(DocumentChunk(
                        id=str(uuid.uuid4()),
                        content=chunk,
                        source_file=file_path.name,
                        page_number=0,
                        chunk_type='text',
                        element_id=f"sekcja_{i+1}"
                    ))
            
            # Wyciągnij obrazy z DOCX
            logger.info("Wyszukiwanie obrazów w DOCX...")
            image_count = 0
            
            # Metoda 1: inline_shapes (obrazy w tekście)
            try:
                for shape_idx, inline_shape in enumerate(doc.inline_shapes):
                    try:
                        # Sprawdź czy to obraz
                        if inline_shape.type == 3:  # 3 = PICTURE
                            logger.debug(f"Znaleziono obraz #{shape_idx + 1} w DOCX")
                            
                            # Wyciągnij dane obrazu
                            image_blob = inline_shape._inline.graphic.graphicData.pic.blipFill.blip.embed
                            image_part = doc.part.related_parts[image_blob]
                            image_bytes = image_part.blob
                            
                            # Zapisz tymczasowo
                            img_path = TEMP_DIR / f"temp_docx_img_{uuid.uuid4()}.png"
                            with open(img_path, 'wb') as f:
                                f.write(image_bytes)
                            
                            # Rozpoznaj przez Gemma 3
                            description = self._describe_image(img_path)
                            if description:
                                logger.debug(f"Wygenerowano opis obrazu z DOCX, długość: {len(description)} znaków")
                                chunks.append(DocumentChunk(
                                    id=str(uuid.uuid4()),
                                    content=description,
                                    source_file=file_path.name,
                                    page_number=0,
                                    chunk_type='image_description',
                                    element_id=f"obraz_docx_{shape_idx + 1}"
                                ))
                                image_count += 1
                            
                            # Usuń temp file
                            img_path.unlink()
                            
                    except Exception as img_error:
                        logger.debug(f"Błąd przetwarzania obrazu {shape_idx + 1} w DOCX: {img_error}")
                        continue
                
                if image_count > 0:
                    logger.info(f"Rozpoznano {image_count} obrazów w DOCX")
                    
            except Exception as e:
                logger.debug(f"Brak obrazów w DOCX lub błąd dostępu: {e}")
            
            logger.info(f"Zakończono przetwarzanie DOCX {file_path}, znaleziono {len(chunks)} fragmentów (w tym {image_count} obrazów)")
        except Exception as e:
            logger.error(f"Błąd podczas przetwarzania DOCX {file_path}: {e}", exc_info=True)
        
        return chunks
    
    def _process_xlsx(self, file_path: Path) -> List[DocumentChunk]:
        """Przetwarza plik XLSX (tekst + obrazy + wykresy)"""
        logger.info(f"Rozpoczynanie przetwarzania XLSX: {file_path}")
        chunks = []
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            logger.info(f"Plik XLSX {file_path} ma {len(workbook.sheetnames)} arkuszy")
            
            total_images = 0
            
            for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]
                sheet_content = []
                
                # Wyciągnij TEKST z komórek
                row_count = 0
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        sheet_content.append(f"Wiersz {row_idx+1}: {row_text}")
                        row_count += 1
                
                if sheet_content:
                    text = f"Arkusz: {sheet_name}\n" + "\n".join(sheet_content)
                    text_chunks = self._chunk_text(text)
                    logger.debug(f"Znaleziono dane w arkuszu {sheet_name} ({row_count} wierszy), podzielono na {len(text_chunks)} fragmentów")
                    for i, chunk in enumerate(text_chunks):
                        chunks.append(DocumentChunk(
                            id=str(uuid.uuid4()),
                            content=chunk,
                            source_file=file_path.name,
                            page_number=0,
                            chunk_type='text',
                            element_id=f"arkusz_{sheet_idx+1}_fragment_{i+1}"
                        ))
                
                # Wyciągnij OBRAZY z arkusza
                logger.info(f"Wyszukiwanie obrazów w arkuszu '{sheet_name}'...")
                try:
                    if hasattr(sheet, '_images') and sheet._images:
                        logger.info(f"Znaleziono {len(sheet._images)} obrazów w arkuszu '{sheet_name}'")
                        
                        for img_idx, image in enumerate(sheet._images):
                            try:
                                logger.debug(f"Przetwarzanie obrazu {img_idx + 1}/{len(sheet._images)} z arkusza '{sheet_name}'")
                                
                                # Wyciągnij dane obrazu (PIL Image)
                                img_path = TEMP_DIR / f"temp_xlsx_img_{uuid.uuid4()}.png"
                                image_pil = image._data()
                                image_pil.save(img_path)
                                
                                # Rozpoznaj przez Gemma 3
                                description = self._describe_image(img_path)
                                if description:
                                    logger.debug(f"Wygenerowano opis obrazu z Excel, długość: {len(description)} znaków")
                                    chunks.append(DocumentChunk(
                                        id=str(uuid.uuid4()),
                                        content=description,
                                        source_file=file_path.name,
                                        page_number=0,
                                        chunk_type='image_description',
                                        element_id=f"obraz_arkusz_{sheet_idx+1}_{img_idx+1}"
                                    ))
                                    total_images += 1
                                
                                # Usuń temp file
                                img_path.unlink()
                                
                            except Exception as img_error:
                                logger.error(f"Błąd przetwarzania obrazu {img_idx + 1} w arkuszu '{sheet_name}': {img_error}")
                                continue
                
                except Exception as e:
                    logger.debug(f"Brak obrazów w arkuszu '{sheet_name}' lub błąd: {e}")
            
            logger.info(f"Zakończono przetwarzanie XLSX {file_path}, znaleziono {len(chunks)} fragmentów (w tym {total_images} obrazów)")
        except Exception as e:
            logger.error(f"Błąd podczas przetwarzania XLSX {file_path}: {e}", exc_info=True)
        
        return chunks
    
    def _process_image(self, file_path: Path) -> List[DocumentChunk]:
        """Przetwarza plik obrazu używając modelu multimodalnego Gemma 3"""
        logger.info(f"Rozpoczynanie przetwarzania obrazu: {file_path}")
        chunks = []
        
        try:
            # Opis obrazu przez Gemma 3 (główna metoda)
            logger.debug("Rozpoczynanie analizy obrazu przez Gemma 3:12B...")
            description = self._describe_image(file_path)
            if description:
                logger.debug(f"Wygenerowano opis grafiki przez Gemma 3, długość: {len(description)} znaków")
                chunks.append(DocumentChunk(
                    id=str(uuid.uuid4()),
                    content=description,
                    source_file=file_path.name,
                    page_number=0,  # Obraz nie ma numeru strony
                    chunk_type='image_description',
                    element_id="opis_grafiki"
                ))
            
            # OCR tekstu z obrazu (opcjonalnie, jeśli Tesseract jest dostępny)
            try:
                logger.debug("Próba OCR tekstu z obrazu (opcjonalnie)...")
                image = Image.open(file_path)
                ocr_text = pytesseract.image_to_string(image, lang='pol')
                
                if ocr_text.strip():
                    text_chunks = self._chunk_text(ocr_text)
                    logger.debug(f"Znaleziono tekst w obrazie przez OCR, podzielono na {len(text_chunks)} fragmentów")
                    for i, chunk in enumerate(text_chunks):
                        chunks.append(DocumentChunk(
                            id=str(uuid.uuid4()),
                            content=chunk,
                            source_file=file_path.name,
                            page_number=0,  # Obraz nie ma numeru strony
                            chunk_type='text',
                            element_id=f"tekst_z_obrazu_{i+1}"
                        ))
            except Exception as ocr_error:
                logger.debug(f"OCR niedostępny (Tesseract nie zainstalowany) - używam tylko Gemma 3: {ocr_error}")
            
            logger.info(f"Zakończono przetwarzanie obrazu {file_path}, znaleziono {len(chunks)} fragmentów")
        except Exception as e:
            logger.error(f"Błąd podczas przetwarzania obrazu {file_path}: {e}", exc_info=True)
        
        return chunks
    
    def _process_audio(self, file_path: Path) -> List[DocumentChunk]:
        """Przetwarza plik audio (transkrypcja + rozpoznawanie mówców)"""
        logger.info(f"Rozpoczynanie przetwarzania pliku audio: {file_path}")
        chunks = []
        
        try:
            # Sprawdź czy whisper jest dostępny
            try:
                import whisper
            except ImportError:
                logger.error("Biblioteka openai-whisper nie jest zainstalowana!")
                logger.error("Zainstaluj: pip install openai-whisper")
                return chunks
            
            # Załaduj model Whisper
            logger.info("=" * 70)
            logger.info("[AUDIO] PRZETWARZANIE PLIKU AUDIO")
            logger.info("=" * 70)
            logger.info(f"Plik: {file_path.name}")
            
            # Sprawdź czy model jest już pobrany
            model_file = WHISPER_MODELS_DIR / "large-v3.pt"

            if not model_file.exists():
                logger.info("[WARNING] MODEL WHISPER NIE JEST POBRANY!")
                logger.info("[DOWNLOAD] Rozpoczynam pobieranie modelu Whisper large-v3 (~3 GB)...")
                logger.info("[INFO] To może potrwać 5-15 minut przy pierwszym użyciu...")
                logger.info(f"[INFO] Model zostanie zapisany w: {model_file.parent}")
                logger.info("[INFO] Kolejne pliki audio będą przetwarzane szybciej (model w cache)")
            else:
                logger.info(f"[OK] Model Whisper large-v3 znaleziony w {model_file.parent}")

            logger.info("[LOADING] Ładowanie modelu Whisper large-v3...")
            load_start = time.time()
            model = whisper.load_model("large-v3", download_root=str(WHISPER_MODELS_DIR))
            load_time = time.time() - load_start
            logger.info(f"[OK] Model Whisper załadowany w {load_time:.2f} sekund")
            
            # Transkrypcja audio
            logger.info(f"Transkrypcja pliku audio: {file_path.name} (może potrwać kilka minut)...")
            start_time = time.time()
            
            result = model.transcribe(
                str(file_path),
                language=None,  # Auto-detect language (lepsze niż wymuszanie "pl")
                task="transcribe",
                verbose=False,
                fp16=False,  # Lepsze dla kompatybilności
                condition_on_previous_text=False  # Lepsze dla krótkich plików
            )
            
            transcription_time = time.time() - start_time
            logger.info(f"Transkrypcja zakończona w {transcription_time:.2f} sekund")
            
            # Pobierz segmenty z timestampami
            segments = result.get("segments", [])
            full_text = result.get("text", "").strip()
            detected_language = result.get("language", "unknown")
            
            logger.info(f"Wykryty język: {detected_language}")
            logger.info(f"Rozpoznano {len(segments)} segmentów audio")
            logger.info(f"Długość transkrypcji: {len(full_text)} znaków")
            
            if len(full_text) > 0 and len(segments) == 0:
                logger.warning("[WARNING] Whisper zwrócił tekst ale bez segmentów - użyję fallback")
            
            # Rozpoznawanie mówców (speaker diarization) - analiza cech audio
            speaker_map = {}
            logger.info("[DIARIZATION] Rozpoznawanie mówców przez analizę barwy głosu...")
            
            try:
                from sklearn.cluster import AgglomerativeClustering
                import numpy as np
                import librosa
                
                # Wczytaj audio
                logger.info(f"[DIARIZATION] Analiza barwy głosu dla {len(segments)} segmentów...")
                audio_data, sr = librosa.load(str(file_path), sr=16000)
                
                # Ekstraktuj cechy audio (MFCC + pitch) dla każdego segmentu
                segment_features = []
                valid_segments = []
                
                for i, seg in enumerate(segments):
                    start_time = seg.get("start", 0)
                    end_time = seg.get("end", 0)
                    
                    # Wyciągnij fragment audio
                    start_sample = int(start_time * sr)
                    end_sample = int(end_time * sr)
                    
                    if end_sample > start_sample and end_sample <= len(audio_data):
                        segment_audio = audio_data[start_sample:end_sample]
                        
                        # Minimum 0.4s do analizy
                        if len(segment_audio) > sr * 0.4:
                            try:
                                # MFCC (Mel-frequency cepstral coefficients) - charakterystyka barwy
                                mfcc = librosa.feature.mfcc(
                                    y=segment_audio, 
                                    sr=sr, 
                                    n_mfcc=13
                                )
                                mfcc_mean = np.mean(mfcc, axis=1)
                                
                                # Pitch (F0) - wysokość głosu
                                pitches, magnitudes = librosa.piptrack(
                                    y=segment_audio, 
                                    sr=sr
                                )
                                pitch_mean = np.mean(pitches[pitches > 0]) if np.any(pitches > 0) else 0
                                
                                # Energy - energia głosu
                                energy = np.mean(librosa.feature.rms(y=segment_audio))
                                
                                # Połącz cechy
                                features = np.concatenate([
                                    mfcc_mean, 
                                    [pitch_mean, energy]
                                ])
                                
                                segment_features.append(features)
                                valid_segments.append(i)
                            except Exception as e:
                                logger.debug(f"Błąd analizy segmentu {i}: {e}")
                
                if len(segment_features) >= 2:
                    # Clustering: grupowanie podobnych głosów
                    features_array = np.array(segment_features)
                    
                    # Normalizacja (ważne dla clustering)
                    from sklearn.preprocessing import StandardScaler
                    scaler = StandardScaler()
                    features_normalized = scaler.fit_transform(features_array)
                    
                    # Hierarchical clustering z automatycznym wykrywaniem liczby klastrów
                    clustering = AgglomerativeClustering(
                        n_clusters=None,
                        distance_threshold=18.0,  # Wysoki threshold = mało klastrów (2-5 osób)
                        linkage='ward'  # Ward minimalizuje wariancję wewnątrz klastrów
                    )
                    labels = clustering.fit_predict(features_normalized)
                    
                    # Mapuj segmenty do mówców
                    for seg_idx, speaker_id in zip(valid_segments, labels):
                        speaker_map[seg_idx] = f"SPEAKER_{speaker_id}"
                    
                    unique_speakers = len(set(labels))
                    logger.info(f"[DIARIZATION] ✅ Wykryto {unique_speakers} mówców (analiza MFCC + pitch)")
                    logger.info(f"[DIARIZATION] Przeanalizowano {len(valid_segments)}/{len(segments)} segmentów")
                else:
                    logger.warning("[DIARIZATION] Za mało segmentów do analizy")
                    raise Exception("Fallback")
                    
            except Exception as diarization_error:
                logger.warning(f"[DIARIZATION] Analiza barwy niedostępna: {diarization_error}")
                logger.info("[DIARIZATION] Fallback: pojedynczy mówca")
                
                # FALLBACK: Zakładamy pojedynczego mówcę
                for i in range(len(segments)):
                    speaker_map[i] = f"SPEAKER_0"
                
                logger.info(f"[DIARIZATION] Fallback: 1 mówca")
            
            # Tworzenie fragmentów z timestampami
            if len(segments) > 0:
                # Normalna ścieżka: używamy segmentów z timestampami
                for i, segment in enumerate(segments):
                    start_time = segment.get("start", 0)
                    end_time = segment.get("end", 0)
                    text = segment.get("text", "").strip()
                    
                    if not text:
                        continue
                    
                    # Formatowanie czasu (MM:SS)
                    start_min = int(start_time // 60)
                    start_sec = int(start_time % 60)
                    end_min = int(end_time // 60)
                    end_sec = int(end_time % 60)
                    
                    # Informacja o mówcy (jeśli dostępna)
                    speaker = speaker_map.get(i, f"SPEAKER_0")  # Fallback: speaker 0
                    
                    # Stwórz fragment z timestampem
                    fragment_text = f"[{start_min:02d}:{start_sec:02d} - {end_min:02d}:{end_sec:02d}] "
                    if speaker:
                        fragment_text += f"[{speaker}] "
                    fragment_text += text
                    
                    chunks.append(DocumentChunk(
                        id=str(uuid.uuid4()),
                        content=fragment_text,
                        source_file=file_path.name,
                        page_number=0,
                        chunk_type='audio_transcription',
                        element_id=f"audio_segment_{i+1}_{start_min:02d}m{start_sec:02d}s"
                    ))
            elif full_text:
                # Fallback: jeśli brak segmentów ale jest pełny tekst, utwórz jeden chunk
                logger.info("[FALLBACK] Brak segmentów, używam pełnego tekstu transkrypcji")
                chunks.append(DocumentChunk(
                    id=str(uuid.uuid4()),
                    content=full_text,
                    source_file=file_path.name,
                    page_number=0,
                    chunk_type='audio_transcription',
                    element_id="audio_full_transcription"
                ))
                logger.info(f"[FALLBACK] Utworzono chunk z pełnym tekstem ({len(full_text)} znaków)")
            
            logger.info(f"Zakończono przetwarzanie audio {file_path.name}, utworzono {len(chunks)} fragmentów")
            
        except Exception as e:
            logger.error(f"Błąd podczas przetwarzania audio {file_path}: {e}", exc_info=True)
        
        return chunks
    
    def _process_video(self, file_path: Path) -> List[DocumentChunk]:
        """Przetwarza plik wideo (audio przez Whisper + klatki przez Gemma 3)"""
        logger.info("=" * 70)
        logger.info("[VIDEO] PRZETWARZANIE PLIKU WIDEO")
        logger.info("=" * 70)
        logger.info(f"Plik: {file_path.name}")
        chunks = []
        
        try:
            # Sprawdź wymagane biblioteki
            try:
                import cv2
                import whisper
            except ImportError as e:
                logger.error(f"Brak wymaganych bibliotek: {e}")
                logger.error("Zainstaluj: pip install opencv-python openai-whisper")
                return chunks
            
            # ===== CZĘŚĆ 1: EKSTRAKCJA AUDIO =====
            logger.info("[STEP 1/3] Ekstrakcja audio z wideo")
            
            # Otwórz wideo
            video = cv2.VideoCapture(str(file_path))
            fps = video.get(cv2.CAP_PROP_FPS)
            total_frames = int(video.get(cv2.CAP_PROP_FRAME_COUNT))
            duration = total_frames / fps if fps > 0 else 0
            
            logger.info(f"[INFO] Parametry wideo:")
            logger.info(f"   FPS: {fps:.2f}")
            logger.info(f"   Klatki: {total_frames}")
            logger.info(f"   Długość: {duration:.2f} sekund ({duration/60:.1f} minut)")
            
            video.release()
            
            # Wyciągnij audio do pliku tymczasowego
            audio_temp = TEMP_DIR / f"temp_audio_{uuid.uuid4()}.wav"
            logger.info(f"[EXTRACT] Ekstrakcja audio do: {audio_temp.name}")
            
            try:
                # Użyj opencv do wyciągnięcia audio
                import subprocess
                result = subprocess.run([
                    'ffmpeg', '-i', str(file_path),
                    '-vn',  # No video
                    '-acodec', 'pcm_s16le',  # WAV format
                    '-ar', '16000',  # 16kHz (Whisper preferuje)
                    '-ac', '1',  # Mono
                    str(audio_temp),
                    '-y'  # Overwrite
                ], capture_output=True, text=True, timeout=300)
                
                if result.returncode != 0 or not audio_temp.exists():
                    logger.warning("ffmpeg niedostępny, próba bez audio...")
                    audio_segments = []
                else:
                    logger.info("[OK] Audio wyekstraktowane")
                    
                    # ===== CZĘŚĆ 2: TRANSKRYPCJA AUDIO =====
                    logger.info("[STEP 2/3] Transkrypcja audio przez Whisper")
                    
                    # Załaduj Whisper
                    logger.info("Ładowanie modelu Whisper large-v3...")
                    whisper_model = whisper.load_model("large-v3")
                    
                    # Transkrypcja
                    logger.info(f"Transkrypcja audio z wideo ({duration:.1f}s)...")
                    transcription_start = time.time()
                    
                    result = whisper_model.transcribe(
                        str(audio_temp),
                        language="pl",
                        task="transcribe",
                        verbose=False
                    )
                    
                    audio_segments = result.get("segments", [])
                    transcription_time = time.time() - transcription_start
                    logger.info(f"[OK] Transkrypcja zakończona w {transcription_time:.2f}s")
                    logger.info(f"   Segmentów audio: {len(audio_segments)}")
                    
                    # Usuń temp audio
                    audio_temp.unlink()
                    
            except Exception as audio_error:
                logger.warning(f"Błąd ekstrakcji/transkrypcji audio: {audio_error}")
                audio_segments = []
            
            # ===== CZĘŚĆ 3: EKSTRAKCJA I ROZPOZNAWANIE KLATEK =====
            logger.info("[STEP 3/3] Ekstrakcja i rozpoznawanie klatek wideo")
            
            video = cv2.VideoCapture(str(file_path))
            
            # Oblicz które klatki wyciągnąć (1 klatka/sekundę)
            frames_to_extract = []
            for second in range(int(duration) + 1):
                frame_num = int(second * fps)
                if frame_num < total_frames:
                    frames_to_extract.append((second, frame_num))
            
            logger.info(f"[FRAMES] Będę analizować {len(frames_to_extract)} klatek (1 klatka/sekundę)")
            
            # Ekstrakcja i rozpoznawanie klatek
            frame_descriptions = {}
            
            for second, frame_num in frames_to_extract:
                try:
                    # Przejdź do klatki
                    video.set(cv2.CAP_PROP_POS_FRAMES, frame_num)
                    ret, frame = video.read()
                    
                    if not ret:
                        continue
                    
                    # Zapisz klatkę tymczasowo
                    frame_temp = TEMP_DIR / f"temp_frame_{uuid.uuid4()}.jpg"
                    cv2.imwrite(str(frame_temp), frame)
                    
                    # Rozpoznaj przez Gemma 3
                    if second % 5 == 0:  # Log co 5 sekund
                        logger.info(f"   Analiza klatki {second}s/{int(duration)}s...")
                    
                    description = self._describe_image(frame_temp)
                    
                    if description:
                        # Usuń prefix "[Opis grafiki]" dla klatek wideo
                        description = description.replace("[Opis grafiki] ", "")
                        frame_descriptions[second] = description
                    
                    # Usuń temp file
                    frame_temp.unlink()
                    
                except Exception as frame_error:
                    logger.debug(f"Błąd przetwarzania klatki {second}s: {frame_error}")
                    continue
            
            video.release()
            logger.info(f"[OK] Rozpoznano {len(frame_descriptions)} klatek wideo")
            
            # ===== CZĘŚĆ 4: ŁĄCZENIE AUDIO + VIDEO =====
            logger.info("[STEP 4/4] Łączenie transkrypcji audio z opisami klatek")
            
            # Stwórz mapę: sekunda → segment audio
            audio_by_second = {}
            for seg in audio_segments:
                start_sec = int(seg.get("start", 0))
                end_sec = int(seg.get("end", 0))
                text = seg.get("text", "").strip()
                
                # Przypisz segment do każdej sekundy którą obejmuje
                for sec in range(start_sec, end_sec + 1):
                    if sec not in audio_by_second:
                        audio_by_second[sec] = []
                    audio_by_second[sec].append(text)
            
            # Tworzenie fragmentów dla każdej sekundy
            for second in range(int(duration) + 1):
                # Audio dla tej sekundy
                audio_text = " ".join(audio_by_second.get(second, ["[cisza]"]))
                
                # Opis klatki dla tej sekundy
                frame_desc = frame_descriptions.get(second, "[brak opisu klatki]")
                
                # Formatuj czas
                minutes = second // 60
                seconds = second % 60
                timestamp = f"{minutes:02d}:{seconds:02d}"
                
                # Połącz audio + video
                fragment_text = f"[{timestamp}]\n"
                fragment_text += f"[Audio] {audio_text}\n"
                fragment_text += f"[Video] {frame_desc}"
                
                chunks.append(DocumentChunk(
                    id=str(uuid.uuid4()),
                    content=fragment_text,
                    source_file=file_path.name,
                    page_number=0,
                    chunk_type='video_transcription',
                    element_id=f"video_second_{second}_{minutes:02d}m{seconds:02d}s"
                ))
            
            logger.info("=" * 70)
            logger.info(f"[OK] ZAKOŃCZONO PRZETWARZANIE WIDEO")
            logger.info(f"   Fragmentów utworzonych: {len(chunks)}")
            logger.info(f"   Audio segmentów: {len(audio_segments)}")
            logger.info(f"   Klatek rozpoznanych: {len(frame_descriptions)}")
            logger.info("=" * 70)
            
        except Exception as e:
            logger.error(f"Błąd podczas przetwarzania wideo {file_path}: {e}", exc_info=True)
        
        return chunks
    
    def _chunk_text(self, text: str, max_chunk_size: int = 500) -> List[str]:
        """Dzieli tekst na fragmenty o maksymalnej długości"""
        if not text.strip():
            return []
        
        # Podział na zdania
        sentences = text.split('. ')
        chunks = []
        current_chunk = ""
        
        for sentence in sentences:
            # Dodajemy kropkę z powrotem (jeśli była)
            sentence = sentence.strip() + ". " if sentence.strip() else ""
            
            # Sprawdzamy czy dodanie zdania nie przekroczy limitu
            if len(current_chunk) + len(sentence) <= max_chunk_size:
                current_chunk += sentence
            else:
                # Jeśli mamy już coś w chunku, zapisujemy i zaczynamy nowy
                if current_chunk.strip():
                    chunks.append(current_chunk.strip())
                current_chunk = sentence
        
        # Dodajemy ostatni chunk jeśli nie jest pusty
        if current_chunk.strip():
            chunks.append(current_chunk.strip())
        
        logger.debug(f"Podzielono tekst na {len(chunks)} fragmentów")
        return chunks
    
    def _describe_image(self, image_path: Path) -> str:
        """Generuje opis obrazu za pomocą modelu multimodalnego (Gemma 3:12B)"""
        logger.debug(f"Rozpoczynanie opisu obrazu: {image_path}")
        start_time = time.time()
        
        try:
            # Konwersja obrazu do base64
            import base64
            with open(image_path, "rb") as image_file:
                encoded_image = base64.b64encode(image_file.read()).decode('utf-8')
            
            logger.debug("Wysyłanie żądania do modelu Gemma 3:12B...")
            # Wysłanie zapytania do Ollama z modelem multimodalnym
            response = requests.post(
                "http://localhost:11434/api/generate",
                json={
                    "model": VISION_MODEL,
                    "prompt": "Opisz szczegółowo co znajduje się na tym obrazie. Odpowiedz po polsku.",
                    "stream": False,
                    "images": [encoded_image]
                },
                timeout=300  # 5 minut timeout dla dużych obrazów
            )
            
            processing_time = time.time() - start_time
            logger.debug(f"Otrzymano odpowiedź od modelu w {processing_time:.2f} sekund")
            
            if response.status_code == 200:
                result = response.json()
                description = result.get('response', '').strip()
                if description:
                    logger.debug(f"Wygenerowano opis grafiki ({len(description)} znaków)")
                    return f"[Opis grafiki] {description}"
                else:
                    logger.warning("Model zwrócił pustą odpowiedź")
                    return ""
            else:
                logger.error(f"Błąd HTTP podczas opisywania obrazu: {response.status_code}")
                return ""
        except requests.exceptions.Timeout:
            logger.error(f"Timeout podczas opisywania obrazu {image_path}")
            return ""
        except Exception as e:
            logger.error(f"Błąd podczas opisywania obrazu {image_path}: {e}", exc_info=True)
            return ""

class EmbeddingProcessor:
    """Klasa do tworzenia embeddingów tekstów"""
    
    def __init__(self, device: str = 'cuda'):
        logger.info(f"Inicjalizacja EmbeddingProcessor (device={device})")
        logger.info("Ładowanie modelu intfloat/multilingual-e5-large...")
        start_time = time.time()
        
        # Inicjalizacja modelu do tworzenia embeddingów
        self.model = SentenceTransformer(
            'intfloat/multilingual-e5-large',
            device=device,
            cache_folder=str(EMBEDDING_MODELS_DIR)
        )
        
        loading_time = time.time() - start_time
        logger.info(f"Załadowano model embeddingów w {loading_time:.2f} sekund")
    
    def create_embeddings(self, chunks: List[DocumentChunk]) -> List[DocumentChunk]:
        """Tworzy embeddingi dla listy fragmentów dokumentów"""
        logger.info(f"Rozpoczynanie tworzenia embeddingów dla {len(chunks)} fragmentów")
        
        if not chunks:
            logger.warning("Brak fragmentów do przetworzenia")
            return chunks
        
        # Przygotowanie tekstów
        texts = [chunk.content for chunk in chunks]
        logger.debug(f"Przygotowano {len(texts)} tekstów do embeddingów")
        
        # Generowanie embeddingów
        logger.info("Rozpoczynanie generowania embeddingów...")
        start_time = time.time()
        
        try:
            # Batch processing dla lepszej wydajności
            batch_size = 32
            embeddings = []
            
            for i in range(0, len(texts), batch_size):
                batch_texts = texts[i:i+batch_size]
                batch_start = time.time()
                
                logger.debug(f"Przetwarzanie batcha {i//batch_size + 1}/{(len(texts)-1)//batch_size + 1} ({len(batch_texts)} tekstów)")
                batch_embeddings = self.model.encode(batch_texts)
                embeddings.extend(batch_embeddings)
                
                batch_time = time.time() - batch_start
                logger.debug(f"Batch {i//batch_size + 1} zakończony w {batch_time:.2f} sekund")
                
                # Logowanie postępu co 5 batchy
                if (i//batch_size + 1) % 5 == 0:
                    processed = min(i + batch_size, len(texts))
                    progress = (processed / len(texts)) * 100
                    logger.info(f"Postęp: {progress:.1f}% ({processed}/{len(texts)})")
            
            # Przypisanie embeddingów do fragmentów
            logger.info("Przypisywanie embeddingów do fragmentów...")
            for i, chunk in enumerate(chunks):
                chunk.embedding = embeddings[i].tolist()
            
            total_time = time.time() - start_time
            logger.info(f"Zakończono tworzenie embeddingów dla {len(chunks)} fragmentów w {total_time:.2f} sekund")
            logger.info(f"Średni czas per fragment: {total_time/len(chunks):.3f} sekund")
            
            # Statystyki długości tekstów
            lengths = [len(text) for text in texts]
            logger.info(f"Statystyki tekstów: min={min(lengths)}, max={max(lengths)}, avg={sum(lengths)/len(lengths):.1f} znaków")
            
        except Exception as e:
            logger.error(f"Błąd podczas tworzenia embeddingów: {e}", exc_info=True)
            raise
        
        return chunks

class VectorDatabase:
    """Klasa do zarządzania bazą wektorową"""
    
    def __init__(self, db_path: str = str(VECTOR_DB_DIR)):
        logger.info(f"Inicjalizacja bazy wektorowej w: {db_path}")
        start_time = time.time()
        
        self.client = chromadb.PersistentClient(
            path=db_path,
            settings=Settings(anonymized_telemetry=False)
        )
        self.collection = self.client.get_or_create_collection("legal_documents")
        
        init_time = time.time() - start_time
        logger.info(f"Baza wektorowa zainicjalizowana w {init_time:.2f} sekund")
    
    def add_documents(self, chunks: List[DocumentChunk]):
        """Dodaje dokumenty do bazy wektorowej"""
        logger.info(f"Rozpoczynanie dodawania {len(chunks)} dokumentów do bazy wektorowej")
        
        if not chunks:
            logger.warning("Brak fragmentów do dodania")
            return
        
        start_time = time.time()
        
        try:
            ids = [chunk.id for chunk in chunks]
            embeddings = [chunk.embedding for chunk in chunks]
            documents = [chunk.content for chunk in chunks]
            metadatas = [
                {
                    "source_file": chunk.source_file,
                    "page_number": chunk.page_number,
                    "chunk_type": chunk.chunk_type,
                    "element_id": chunk.element_id
                }
                for chunk in chunks
            ]
            
            logger.debug("Wysyłanie danych do bazy wektorowej...")
            self.collection.add(
                ids=ids,
                embeddings=embeddings,
                documents=documents,
                metadatas=metadatas
            )
            
            total_time = time.time() - start_time
            logger.info(f"Zakończono dodawanie dokumentów do bazy w {total_time:.2f} sekund")
            logger.info(f"Średni czas per dokument: {total_time/len(chunks):.3f} sekund")
            
        except Exception as e:
            logger.error(f"Błąd podczas dodawania dokumentów do bazy: {e}", exc_info=True)
            raise
    
    def search(self, query: str, n_results: int = 5) -> List[SourceReference]:
        """Wyszukuje dokumenty pasujące do zapytania"""
        logger.info(f"Rozpoczynanie wyszukiwania dla zapytania: {query}")
        start_time = time.time()
        
        try:
            # Utworzenie embeddingu dla zapytania
            logger.debug("Tworzenie embeddingu dla zapytania...")
            model = SentenceTransformer(
                'intfloat/multilingual-e5-large',
                cache_folder=str(EMBEDDING_MODELS_DIR)
            )
            query_embedding = model.encode([query]).tolist()
            
            # Wyszukiwanie
            logger.debug("Wyszukiwanie w bazie wektorowej...")
            results = self.collection.query(
                query_embeddings=query_embedding,
                n_results=n_results
            )
            
            # Formatowanie wyników
            source_references = []
            for i in range(len(results['ids'][0])):
                source_references.append(SourceReference(
                    source_file=results['metadatas'][0][i]['source_file'],
                    page_number=results['metadatas'][0][i]['page_number'],
                    element_id=results['metadatas'][0][i]['element_id'],
                    content=results['documents'][0][i],
                    distance=results['distances'][0][i] if 'distances' in results else 0.0
                ))
            
            search_time = time.time() - start_time
            logger.info(f"Zakończono wyszukiwanie w {search_time:.2f} sekund, znaleziono {len(source_references)} wyników")
            
            return source_references
            
        except Exception as e:
            logger.error(f"Błąd podczas wyszukiwania: {e}", exc_info=True)
            return []

class RAGSystem:
    """Główna klasa systemu RAG"""
    
    def __init__(self, config_file: str = "auth_config.json", device_mode: str = 'auto'):
        logger.info("Inicjalizacja systemu RAG")
        
        # Inicjalizacja Device Manager
        self.device_manager = DeviceManager(mode=device_mode)
        logger.info(f"Device configuration: {self.device_manager.config}")
        
        # Komponenty z device assignment
        self.doc_processor = DocumentProcessor()
        embeddings_device = self.device_manager.get_device('embeddings')
        self.embedding_processor = EmbeddingProcessor(device=embeddings_device)
        self.vector_db = VectorDatabase()
        self.greeting_filter = GreetingFilter()  # Filtr powitań
        
        # Inicjalizacja Model Provider (OpenAI lub Ollama)
        self.config = self._load_config(config_file)
        self.model_provider = self._initialize_model_provider()
        
        # Inicjalizacja Hybrydowego Wyszukiwania
        reranker_device = self.device_manager.get_device('reranker')
        self.hybrid_search = self._initialize_hybrid_search(reranker_device=reranker_device)
        
        # Inicjalizacja Audit Logger
        self.audit_logger = get_audit_logger()
        
        # Inicjalizacja Web Search (opcjonalne)
        bing_search, web_scraper, search_cache = self._initialize_web_search()
        self.bing_search = bing_search
        self.web_scraper = web_scraper
        self.search_cache = search_cache
        
        logger.info("System RAG zainicjalizowany")
    
    def _load_config(self, config_file: str) -> Dict[str, Any]:
        """Wczytuje konfigurację z pliku JSON"""
        config_path = BASE_DIR / config_file
        try:
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                logger.info(f"Wczytano konfigurację z {config_file}")
                return config
            else:
                logger.warning(f"Brak pliku {config_file}, używam domyślnej konfiguracji")
                return {}
        except Exception as e:
            logger.error(f"Błąd podczas wczytywania konfiguracji: {e}")
            return {}
    
    def _initialize_model_provider(self) -> ModelProvider:
        """Inicjalizuje odpowiedni provider modelu (OpenAI lub Ollama)"""
        try:
            # Przygotuj konfigurację dla ModelFactory
            provider_config = {}
            
            # OpenAI config
            if 'openai' in self.config:
                openai_cfg = self.config['openai']
                provider_config['openai_api_key'] = openai_cfg.get('api_key', '')
                provider_config['openai_model'] = openai_cfg.get('model', None)
            
            # Ollama config (fallback)
            if 'ollama' in self.config:
                ollama_cfg = self.config['ollama']
                provider_config['ollama_model'] = ollama_cfg.get('model', 'gemma3:12b')
                provider_config['ollama_url'] = ollama_cfg.get('url', 'http://127.0.0.1:11434')
            else:
                # Domyślne wartości jeśli brak w config
                provider_config['ollama_model'] = 'gemma3:12b'
                provider_config['ollama_url'] = 'http://127.0.0.1:11434'
            
            # Utwórz provider
            provider = ModelFactory.create_provider(provider_config)
            logger.info(f"Model provider: {provider.__class__.__name__} - {provider.get_model_name()}")
            
            return provider
            
        except Exception as e:
            logger.error(f"Błąd podczas inicjalizacji model provider: {e}")
            # Fallback: spróbuj Ollama z domyślnymi ustawieniami
            logger.info("Próba fallback na Ollama z domyślnymi ustawieniami...")
            from model_provider import OllamaProvider
            return OllamaProvider()
    
    def _initialize_hybrid_search(self, reranker_device: str = 'cuda') -> HybridSearch:
        """Inicjalizuje hybrydowe wyszukiwanie (Vector + BM25 + Reranking)"""
        try:
            logger.info(f"Hybrydowe wyszukiwanie: reranker device = {reranker_device}")
            
            # Utwórz HybridSearch
            hybrid_search = HybridSearch(
                vector_db=self.vector_db,
                cache_dir=VECTOR_DB_DIR,
                use_bm25=True,
                use_reranker=True,
                reranker_device=reranker_device
            )
            
            logger.info("Hybrydowe wyszukiwanie zainicjalizowane")
            return hybrid_search
            
        except Exception as e:
            logger.error(f"Błąd podczas inicjalizacji hybrydowego wyszukiwania: {e}")
            logger.warning("Fallback: używam tylko vector search")
            # Zwróć None - będziemy używać standardowego vector search
            return None
    
    def _initialize_web_search(self) -> tuple:
        """Inicjalizuje web search (Bing API + Scraper)"""
        try:
            web_config = self.config.get('web_search', {})
            
            if not web_config.get('enabled', False):
                logger.info("Web search wyłączone w konfiguracji")
                return None, None, None
            
            bing_key = web_config.get('bing_api_key', '').strip()
            
            if not bing_key:
                logger.warning("Brak Bing API key, web search niedostępne")
                return None, None, None
            
            # Inicjalizuj komponenty
            bing_search = BingSearchProvider(api_key=bing_key)
            web_scraper = WebScraper()
            cache_ttl = web_config.get('cache_ttl_hours', 24)
            search_cache = WebSearchCache(ttl_hours=cache_ttl)
            
            logger.info("Web search zainicjalizowany (Bing API + Scraper + Cache)")
            return bing_search, web_scraper, search_cache
            
        except Exception as e:
            logger.error(f"Błąd podczas inicjalizacji web search: {e}")
            return None, None, None
    
    def rebuild_bm25_index(self):
        """Przebudowuje BM25 index (wywołaj po dodaniu nowych dokumentów)"""
        if self.hybrid_search:
            try:
                logger.info("Przebudowywanie BM25 index...")
                self.hybrid_search.build_bm25_index()
                logger.info("BM25 index przebudowany")
            except Exception as e:
                logger.error(f"Błąd podczas przebudowy BM25 index: {e}")
        else:
            logger.warning("Hybrydowe wyszukiwanie nie jest dostępne")
    
    def index_documents(self, data_directory: str):
        """Indeksuje dokumenty z katalogu"""
        logger.info("="*60)
        logger.info("ROZPOCZYNAM INDEKSOWANIE DOKUMENTÓW")
        logger.info("="*60)
        start_time = time.time()
        
        try:
            # Przetwarzanie dokumentów
            logger.info("Etap 1: Przetwarzanie dokumentów")
            chunks = self.doc_processor.process_directory(data_directory)
            logger.info(f"Etap 1 zakończony: {len(chunks)} fragmentów dokumentów")
            
            if not chunks:
                logger.warning("Nie znaleziono żadnych fragmentów do indeksowania")
                return
            
            # Tworzenie embeddingów
            logger.info("Etap 2: Tworzenie embeddingów")
            chunks_with_embeddings = self.embedding_processor.create_embeddings(chunks)
            logger.info("Etap 2 zakończony: Tworzenie embeddingów")
            
            # Zapis do bazy wektorowej
            logger.info("Etap 3: Zapis do bazy wektorowej")
            self.vector_db.add_documents(chunks_with_embeddings)
            logger.info("Etap 3 zakończony: Zapis do bazy wektorowej")
            
            total_time = time.time() - start_time
            logger.info("="*60)
            logger.info(f"INDEKSOWANIE ZAKOŃCZONE POMYŚLNIE W {total_time:.2f} SEKUND")
            logger.info(f"Przetworzono {len(chunks)} fragmentów dokumentów")
            logger.info("="*60)
            
            # Przebuduj BM25 index dla hybrydowego wyszukiwania
            logger.info("Przebudowywanie BM25 index...")
            self.rebuild_bm25_index()
            
        except Exception as e:
            logger.error(f"Błąd podczas indeksowania dokumentów: {e}", exc_info=True)
            raise
    
    def _format_source_info(self, source: SourceReference) -> str:
        """Formatuje informacje o źródle"""
        info_parts = [f"Dokument: {source.source_file}"]
        
        if source.page_number > 0:
            info_parts.append(f"Strona: {source.page_number}")
        
        if source.element_id:
            info_parts.append(f"Element: {source.element_id}")
            
        return ", ".join(info_parts)
    
    def query(self, question: str, n_results: int = 3, user_id: str = 'anonymous', session_id: str = None, 
              temperature: float = 0.1, top_p: float = 0.85, top_k: int = 30, max_tokens: int = 1000) -> str:
        """Odpowiada na pytanie użytkownika"""
        logger.info("="*60)
        logger.info(f"ROZPOCZYNAM ODPOWIADANIE NA PYTANIE: {question}")
        logger.info("="*60)
        start_time = time.time()
        
        # Generate session_id if not provided
        if session_id is None:
            import uuid
            session_id = str(uuid.uuid4())[:8]
        
        try:
            # FILTROWANIE POWITAŃ (nowe!)
            original_question = question
            question_cleaned, had_greeting, _ = self.greeting_filter.filter_with_info(question)
            
            if had_greeting:
                logger.info(f"Usunięto powitanie z pytania: '{original_question}' -> '{question_cleaned}'")
            
            # Sprawdź czy po filtrowaniu zostało jakieś pytanie
            if not question_cleaned or len(question_cleaned) < 3:
                logger.warning("Pytanie jest puste po usunięciu powitań")
                return "Proszę zadaj pytanie dotyczące dokumentów w bazie."
            
            # Używamy oczyszczonego pytania
            question = question_cleaned
            
            # Wyszukiwanie pasujących dokumentów (HYBRYDOWE: Vector + BM25 + Reranking)
            logger.info("Etap 1: Hybrydowe wyszukiwanie pasujących dokumentów")
            
            if self.hybrid_search:
                # Użyj hybrydowego wyszukiwania
                try:
                    hybrid_results = self.hybrid_search.search(question, top_k=n_results)
                    
                    # Konwertuj do formatu SourceReference
                    results = []
                    for doc in hybrid_results:
                        # Pobierz metadane - mogą być w różnych formatach
                        # UWAGA: doc może zawierać 'id', które NIE jest używane w SourceReference
                        metadata = doc.get('metadata', {})
                        if isinstance(metadata, dict):
                            source_file = metadata.get('source_file', '')
                            page_number = metadata.get('page_number', metadata.get('page', 0))
                            chunk_type = metadata.get('chunk_type', 'text')
                            element_id = metadata.get('element_id', '')
                        else:
                            # Fallback jeśli metadata nie jest dict
                            source_file = ''
                            page_number = 0
                            chunk_type = 'text'
                            element_id = ''
                        
                        # Konwersja score na distance (rerank_score lub rrf_score)
                        score = doc.get('rerank_score', doc.get('rrf_score', 0.5))
                        distance = 1.0 - score if score <= 1.0 else 1.0 / (1.0 + score)  # Normalizacja jeśli score > 1
                        
                        # Tworzenie SourceReference - NIE przekazujemy 'id' z doc
                        results.append(SourceReference(
                            content=doc.get('content', ''),
                            source_file=source_file,
                            page_number=page_number,
                            element_id=element_id,
                            distance=distance
                        ))
                    
                    logger.info(f"Hybrydowe wyszukiwanie: znaleziono {len(results)} dokumentów")
                    
                except Exception as e:
                    logger.error(f"Błąd hybrydowego wyszukiwania: {e}", exc_info=True)
                    logger.info("Fallback: używam prostego vector search")
                    results = self.vector_db.search(question, n_results)
            else:
                # Fallback: prosty vector search
                logger.info("Używam prostego vector search (hybrid search niedostępny)")
                results = self.vector_db.search(question, n_results)
            
            if not results:
                logger.warning("Nie znaleziono odpowiednich informacji w bazie danych")
                return "Nie znaleziono odpowiednich informacji w bazie danych."
            
            logger.info(f"Etap 1 zakończony: Znaleziono {len(results)} pasujących dokumentów")
            
            # Przygotowanie kontekstu dla modelu
            logger.info("Etap 2: Przygotowanie kontekstu dla modelu")
            context_parts = []
            sources_info = []
            
            for i, result in enumerate(results):
                source_info = self._format_source_info(result)
                context_parts.append(f"[{i+1}] {source_info}\nFragment: {result.content}")
                sources_info.append(source_info)
            
            context = "\n\n".join(context_parts)
            logger.info("Etap 2 zakończony: Przygotowano kontekst")
            
            # Przygotowanie promptu dla modelu
            system_prompt = """Jesteś asystentem analizującym dokumenty. Twoim zadaniem jest odpowiedzieć na pytanie użytkownika WYŁĄCZNIE na podstawie dostarczonych fragmentów dokumentów.

WAŻNE ZASADY:
1. Odpowiadaj TYLKO na podstawie informacji zawartych w dostarczonych fragmentach dokumentów
2. NIE używaj swojej ogólnej wiedzy ani informacji spoza dostarczonych dokumentów
3. Jeśli informacji nie ma w dostarczonych fragmentach, napisz: "Nie znalazłem informacji na ten temat w dostarczonych dokumentach"
4. Podsumuj znalezione informacje i wyjaśnij co one znaczą
5. Zawsze wskazuj z którego fragmentu pochodzi informacja (np. "[1]", "[2]")
6. Odpowiadaj w języku polskim"""
            
            user_context = f"""Pytanie: {question}

Fragmenty dokumentów:
{context}

Odpowiedź (bazując TYLKO na powyższych fragmentach):"""
            
            # Wysłanie zapytania do modelu (OpenAI lub Ollama)
            logger.info(f"Etap 3: Generowanie odpowiedzi przez model {self.model_provider.get_model_name()}")
            response_start = time.time()
            
            try:
                answer = self.model_provider.generate(
                    prompt=system_prompt,
                    context=user_context,
                    temperature=temperature,
                    max_tokens=max_tokens,
                    top_k=top_k,
                    top_p=top_p,
                    timeout=300
                )
                
                response_time = time.time() - response_start
                logger.info(f"Etap 3 zakończony: Odpowiedź wygenerowana w {response_time:.2f} sekund")
                
                # Dodanie informacji o źródłach
                sources_text = "\n\nŹródła:\n" + "\n".join([
                    f"[{i+1}] {info}"
                    for i, info in enumerate(sources_info)
                ])
                
                total_time = time.time() - start_time
                logger.info("="*60)
                logger.info(f"ODPOWIEDŹ WYGENEROWANA POMYŚLNIE W {total_time:.2f} SEKUND")
                logger.info("="*60)
                
                # AUDIT LOG (nowe!)
                try:
                    # Przygotuj źródła dla audit log
                    audit_sources = [
                        {
                            'source_file': result.source_file,
                            'page': result.page_number,
                            'element_id': result.element_id,
                            'chunk_type': result.chunk_type
                        }
                        for result in results
                    ]
                    
                    # Loguj zapytanie
                    self.audit_logger.log_query(
                        user_id=user_id,
                        session_id=session_id,
                        query=original_question,  # Użyj oryginalnego pytania (przed filtrowaniem)
                        response=answer,
                        sources=audit_sources,
                        model=self.model_provider.get_model_name(),
                        time_ms=total_time * 1000
                    )
                except Exception as audit_error:
                    logger.warning(f"Błąd audit log: {audit_error}")
                
                return answer + sources_text
                
            except Exception as model_error:
                logger.error(f"Błąd podczas generowania odpowiedzi przez model: {model_error}", exc_info=True)
                return f"Wystąpił błąd podczas generowania odpowiedzi: {str(model_error)}"
                
        except Exception as e:
            logger.error(f"Błąd podczas przetwarzania zapytania: {e}", exc_info=True)
            return "Wystąpił błąd podczas przetwarzania zapytania."
    
    def generate_questions_for_file(self, file_name: str, max_questions: int = 3) -> List[str]:
        """Generuje przykładowe pytania dla danego pliku na podstawie jego treści"""
        logger.info(f"Generowanie pytań dla pliku: {file_name}")
        
        try:
            # Pobierz fragmenty z tego pliku
            results = self.vector_db.collection.get(
                where={"source_file": file_name},
                include=['documents', 'metadatas'],
                limit=10  # Weź pierwsze 10 fragmentów
            )
            
            if not results['documents']:
                logger.warning(f"Brak fragmentów dla pliku: {file_name}")
                return []
            
            # Połącz fragmenty w jeden kontekst
            content_samples = results['documents'][:5]  # Pierwsze 5 fragmentów
            context = "\n\n".join(content_samples)
            
            # Ogranicz długość kontekstu
            if len(context) > 2000:
                context = context[:2000] + "..."
            
            # Określ typ dokumentu
            file_ext = Path(file_name).suffix.lower()
            doc_type = "obraz" if file_ext in ['.jpg', '.jpeg', '.png', '.bmp'] else "dokument"
            
            # Prompt dla modelu
            system_prompt = f"""Wygeneruj {max_questions} konkretne pytania na podstawie podanego fragmentu dokumentu.

Zasady:
1. Pytania muszą być konkretne i szczegółowe
2. Odpowiedź na każde pytanie MUSI być w podanym fragmencie
3. Pytania powinny być praktyczne i użyteczne
4. {'Dla obrazów: pytaj o widoczne elementy, obiekty, kolory, kompozycję' if doc_type == 'obraz' else 'Dla dokumentów: pytaj o fakty, zasady, procedury, definicje'}
5. Odpowiedz TYLKO listą pytań, każde w nowej linii, bez numeracji"""
            
            user_context = f"""Plik: {file_name}

Fragment z dokumentu:
{context}

Pytania:"""
            
            # Wywołaj model
            try:
                generated_text = self.model_provider.generate(
                    prompt=system_prompt,
                    context=user_context,
                    temperature=0.7,
                    max_tokens=500,
                    timeout=120
                )
                
                # Parsuj pytania (każde w nowej linii)
                questions = []
                for line in generated_text.split('\n'):
                    line = line.strip()
                    # Usuń numerację jeśli jest
                    line = line.lstrip('0123456789.-) ')
                    if line and len(line) > 10 and '?' in line:
                        questions.append(line)
                
                # Ogranicz do max_questions
                questions = questions[:max_questions]
                
                logger.info(f"Wygenerowano {len(questions)} pytań dla {file_name}")
                return questions
                
            except Exception as model_error:
                logger.error(f"Błąd podczas generowania pytań przez model: {model_error}")
                return []
                
        except Exception as e:
            logger.error(f"Błąd podczas generowania pytań: {e}", exc_info=True)
            return []

def load_suggested_questions() -> List[dict]:
    """Wczytuje sugerowane pytania z pliku"""
    if SUGGESTED_QUESTIONS_FILE.exists():
        try:
            with open(SUGGESTED_QUESTIONS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Błąd wczytywania pytań: {e}")
            return []
    return []

def save_suggested_questions(questions: List[dict]):
    """Zapisuje sugerowane pytania do pliku"""
    try:
        with open(SUGGESTED_QUESTIONS_FILE, 'w', encoding='utf-8') as f:
            json.dump(questions, f, ensure_ascii=False, indent=2)
        logger.info(f"Zapisano {len(questions)} pytań do {SUGGESTED_QUESTIONS_FILE}")
    except Exception as e:
        logger.error(f"Błąd zapisywania pytań: {e}")

def add_questions_for_file(file_name: str, rag_system: 'RAGSystem', max_questions: int = 3):
    """Dodaje pytania dla nowego pliku do listy sugerowanych"""
    logger.info(f"Dodawanie pytań dla nowego pliku: {file_name}")
    
    # Wygeneruj pytania
    new_questions = rag_system.generate_questions_for_file(file_name, max_questions)
    
    if not new_questions:
        return
    
    # Wczytaj istniejące pytania
    all_questions = load_suggested_questions()
    
    # Dodaj nowe pytania
    for question in new_questions:
        all_questions.append({
            "question": question,
            "source_file": file_name,
            "generated_at": time.strftime("%Y-%m-%d %H:%M:%S")
        })
    
    # Ogranicz do 30 najnowszych
    all_questions = all_questions[-30:]
    
    # Zapisz
    save_suggested_questions(all_questions)
    logger.info(f"Dodano {len(new_questions)} pytań. Łącznie: {len(all_questions)}")

def main():
    """Główna funkcja programu"""
    # Konfiguracja logowania dla głównej funkcji
    main_logger = logging.getLogger("main")
    main_logger.info("Uruchamianie systemu RAG")
    
    # Inicjalizacja systemu
    rag_system = RAGSystem()
    
    # Sprawdzenie argumentów linii poleceń
    if len(sys.argv) < 2:
        print("Użycie:")
        print("  Indeksowanie dokumentów: python rag_system.py index <ścieżka_do_katalogu>")
        print("  Zadanie pytania: python rag_system.py query \"Twoje pytanie\"")
        return
    
    command = sys.argv[1]
    
    if command == "index":
        if len(sys.argv) < 3:
            print("Podaj ścieżkę do katalogu z dokumentami")
            return
        
        data_directory = sys.argv[2]
        rag_system.index_documents(data_directory)
    
    elif command == "query":
        if len(sys.argv) < 3:
            print("Podaj pytanie")
            return
        
        question = " ".join(sys.argv[2:])
        answer = rag_system.query(question)
        print("\nOdpowiedź:")
        print(answer)
    
    else:
        print("Nieznana komenda. Dostępne komendy: index, query")

if __name__ == "__main__":
    main()